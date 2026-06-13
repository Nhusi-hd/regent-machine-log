"""
excel_monthly.py
Xuất file Excel tháng — đúng format template gốc:
  - Mỗi dự án = 1 sheet
  - Rows = các chỉ số (Overall + Defect)
  - Columns = các ngày làm việc trong tháng
  - Màu vàng = nhập tay, màu xanh = công thức
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
from datetime import date, timedelta

# ── Colors ────────────────────────────────────────────────────────
NAVY       = "1A2744"
ORANGE     = "E8542A"
YELLOW     = "FFD700"   # nhập tay
BLUE       = "00B0F0"   # công thức
GREEN_HDR  = "00B050"   # section header
WHITE      = "FFFFFF"
GRAY       = "F2F2F2"
LIGHT_BLUE = "DDEEFF"

def _fill(c): return PatternFill("solid", fgColor=c, start_color=c)
def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _cell(ws, r, c, val=None, fill=None, bold=False,
          fcolor="000000", size=10, align="center", wrap=False, border=True):
    cell = ws.cell(row=r, column=c, value=val)
    if fill:  cell.fill      = _fill(fill)
    cell.font      = _font(bold=bold, color=fcolor, size=size)
    cell.alignment = _align(align, "center", wrap)
    if border: cell.border  = _border()
    return cell


# ── Row definitions (matching template) ──────────────────────────
# (label, db_field_or_None, color, is_formula, format)
OVERALL_ROWS = [
    ("Tên Dự Án",                "project",               GREEN_HDR, False, "text"),
    ("MO/Color",                  "mo_color",              YELLOW,    False, "text"),
    ("SAM (testing)",             "sam",                   YELLOW,    False, "num2"),
    ("Target output / day (pcs)", "target_output_day",     BLUE,      True,  "int"),
    ("Auto target output/h/HC",   "auto_target_h_hc",      BLUE,      True,  "num2"),
    ("Actual output/day/HC(pcs)", "actual_output_day_hc",  YELLOW,    False, "int"),
    ("Actual output/day/MC(pcs)", "actual_output_day_mc",  YELLOW,    False, "int"),
    ("Auto output/h/HC (pcs)",    "auto_output_h_hc",      YELLOW,    False, "num2"),
    ("Actual output/h/MC (pcs)",  "actual_output_h_mc",    YELLOW,    False, "num2"),
    ("Defect rate (%)",           "defect_rate_pct",       BLUE,      True,  "pct"),
    ("EFF(%)",                    "eff_pct",               BLUE,      True,  "pct"),
    ("Manual output/day/HC (pcs)","manual_output_day_hc",  YELLOW,    False, "int"),
    ("Manual output/h/HC (pcs)", "manual_output_h_hc",    YELLOW,    False, "num2"),
    ("Machine working time (min)","machine_working_time",  YELLOW,    False, "int"),
    ("Change-over time(min)",     "changeover_time",       YELLOW,    False, "int"),
    ("Breakdown time(min)",       "breakdown_time",        YELLOW,    False, "int"),
    ("Idle time(min)",            "idle_time",             YELLOW,    False, "int"),
    ("MC Utilization (%)",        "mc_utilization_pct",    BLUE,      True,  "pct"),
]

def _get_working_days(year: int, month: int) -> list:
    """Trả về list các date là ngày làm việc (Mon–Sat) trong tháng."""
    _, last = calendar.monthrange(year, month)
    days = []
    for d in range(1, last + 1):
        dt = date(year, month, d)
        if dt.weekday() < 6:   # 0=Mon … 5=Sat, 6=Sun
            days.append(dt)
    return days


def build_monthly_excel(
    records: list,          # list of dicts from DB
    year: int,
    month: int,
    output_path: str,
):
    wb = Workbook()
    wb.remove(wb.active)    # xóa sheet mặc định

    # ── Nhóm records theo project ──────────────────────────────────
    proj_map: dict = {}
    for r in records:
        proj_map.setdefault(r["project"], []).append(r)

    if not proj_map:
        # Tạo sheet rỗng nếu không có data
        ws = wb.create_sheet("No Data")
        ws["A1"] = f"Không có dữ liệu tháng {month}/{year}"
        wb.save(output_path)
        return

    working_days = _get_working_days(year, month)
    month_name   = f"{calendar.month_abbr[month]}-{year}"

    for proj_name, proj_records in proj_map.items():
        # Map date → record cho project này
        date_map = {r["date"]: r for r in proj_records}

        # Sheet name tối đa 31 ký tự
        sheet_name = proj_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # ── Column widths ─────────────────────────────────────────
        ws.column_dimensions["A"].width = 12   # section label
        ws.column_dimensions["B"].width = 32   # row label
        for i, _ in enumerate(working_days, start=3):
            ws.column_dimensions[get_column_letter(i)].width = 11

        # ── Title row ─────────────────────────────────────────────
        last_col = 2 + len(working_days)
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=last_col)
        title = ws.cell(row=1, column=1,
            value=f"DAILY MACHINE PERFORMANCE LOG — {proj_name.upper()} — {month_name}")
        title.fill      = _fill(NAVY)
        title.font      = _font(bold=True, color=WHITE, size=12)
        title.alignment = _align("center", "center")
        title.border    = _border()
        ws.row_dimensions[1].height = 28

        # ── Date header row ───────────────────────────────────────
        ws.row_dimensions[2].height = 20
        _cell(ws, 2, 1, "Overall",    NAVY, bold=True, fcolor=WHITE, size=10)
        _cell(ws, 2, 2, "Date",       NAVY, bold=True, fcolor=WHITE, size=10)
        for i, day in enumerate(working_days, start=3):
            label = f"{day.day:02d}-{calendar.month_abbr[month]}"
            c = _cell(ws, 2, i, label, NAVY, bold=True, fcolor=WHITE, size=10)

        # ── Overall rows ──────────────────────────────────────────
        ROW_START = 3
        for row_idx, (label, field, color, is_formula, fmt) in enumerate(OVERALL_ROWS):
            r = ROW_START + row_idx
            ws.row_dimensions[r].height = 16

            # Section label (cột A) — chỉ viết "Overall" ở dòng đầu
            if row_idx == 0:
                ws.merge_cells(start_row=r, start_column=1,
                               end_row=r + len(OVERALL_ROWS) - 1, end_column=1)
                sec = ws.cell(row=r, column=1, value="Overall")
                sec.fill      = _fill(GRAY)
                sec.font      = _font(bold=True, size=10)
                sec.alignment = _align("center", "center")
                sec.border    = _border()

            # Row label (cột B)
            lbl_fill = color if row_idx == 0 else (BLUE if is_formula else YELLOW)
            lbl_fcolor = WHITE if row_idx == 0 else ("0070C0" if is_formula else "000000")
            _cell(ws, r, 2, label, lbl_fill, bold=is_formula,
                  fcolor=lbl_fcolor, align="left")

            # Data cells — mỗi ngày 1 cột
            for col_idx, day in enumerate(working_days, start=3):
                date_str = day.strftime("%Y-%m-%d")
                rec      = date_map.get(date_str)
                val      = None
                if rec and field:
                    val = rec.get(field)

                # Format value
                if val is not None:
                    if fmt == "int":    val = int(val)
                    elif fmt == "num2": val = round(float(val), 2)
                    elif fmt == "pct":  val = round(float(val), 1)

                cell_fill  = BLUE if is_formula else YELLOW
                cell_fc    = "0070C0" if is_formula else "000000"
                cell_bold  = is_formula
                c = _cell(ws, r, col_idx, val, cell_fill,
                          bold=cell_bold, fcolor=cell_fc)

                # Number format
                if fmt == "pct" and val is not None:
                    c.number_format = '0.0"%"'
                elif fmt == "int" and val is not None:
                    c.number_format = '#,##0'

        # ── Defect section ────────────────────────────────────────
        # Thu thập tất cả tên lỗi xuất hiện trong tháng
        defect_names: list = []
        for rec in proj_records:
            for d in (rec.get("defects") or []):
                if d.get("name") and d["name"] not in defect_names:
                    defect_names.append(d["name"])

        DEFECT_ROW_START = ROW_START + len(OVERALL_ROWS) + 1
        ws.row_dimensions[DEFECT_ROW_START - 1].height = 6  # spacer

        # Section header "Defect (pcs)"
        ws.merge_cells(start_row=DEFECT_ROW_START, start_column=1,
                       end_row=DEFECT_ROW_START,   end_column=2)
        dh = ws.cell(row=DEFECT_ROW_START, column=1, value="Defect (pcs)")
        dh.fill      = _fill(NAVY)
        dh.font      = _font(bold=True, color=WHITE, size=10)
        dh.alignment = _align("center")
        dh.border    = _border()
        _cell(ws, DEFECT_ROW_START, 2, None, NAVY)

        # Column headers lặp lại cho section defect
        for col_idx, day in enumerate(working_days, start=3):
            label = f"{day.day:02d}-{calendar.month_abbr[month]}"
            _cell(ws, DEFECT_ROW_START, col_idx, label,
                  NAVY, bold=True, fcolor=WHITE, size=10)
        ws.row_dimensions[DEFECT_ROW_START].height = 18

        # Defect name rows
        for di, dname in enumerate(defect_names):
            r = DEFECT_ROW_START + 1 + di
            ws.row_dimensions[r].height = 15
            _cell(ws, r, 1, "Defect", GRAY, bold=False, align="center")
            _cell(ws, r, 2, dname, YELLOW, align="left")
            for col_idx, day in enumerate(working_days, start=3):
                date_str = day.strftime("%Y-%m-%d")
                rec      = date_map.get(date_str)
                qty      = 0
                if rec:
                    for d in (rec.get("defects") or []):
                        if d.get("name") == dname:
                            qty = int(d.get("qty", 0))
                            break
                _cell(ws, r, col_idx, qty if qty else None, YELLOW)

        # Total defect row
        total_r = DEFECT_ROW_START + 1 + len(defect_names)
        ws.row_dimensions[total_r].height = 16
        ws.merge_cells(start_row=total_r, start_column=1,
                       end_row=total_r,   end_column=2)
        td = ws.cell(row=total_r, column=1, value="Total defect")
        td.fill      = _fill(BLUE)
        td.font      = _font(bold=True, color="0070C0", size=10)
        td.alignment = _align("center")
        td.border    = _border()
        _cell(ws, total_r, 2, None, BLUE)
        for col_idx, day in enumerate(working_days, start=3):
            date_str = day.strftime("%Y-%m-%d")
            rec      = date_map.get(date_str)
            tot      = int(rec.get("total_defect", 0)) if rec else None
            _cell(ws, total_r, col_idx, tot, BLUE,
                  bold=True, fcolor="0070C0")

        # ── Legend + Footer ───────────────────────────────────────
        foot_r = total_r + 2
        ws.merge_cells(start_row=foot_r, start_column=1,
                       end_row=foot_r,   end_column=last_col)
        leg = ws.cell(row=foot_r, column=1,
            value="⬛ Vàng = Nhập tay     🔵 Xanh = Công thức tự tính")
        leg.fill      = _fill(GRAY)
        leg.font      = _font(size=8, color="666666")
        leg.alignment = _align("left")

        ws.merge_cells(start_row=foot_r+1, start_column=1,
                       end_row=foot_r+1,   end_column=last_col)
        ft = ws.cell(row=foot_r+1, column=1,
            value="Regent Garment Factory  |  Crystal International Group  |  Innovation Engineering Dept.")
        ft.fill      = _fill(NAVY)
        ft.font      = _font(bold=False, color=WHITE, size=8)
        ft.alignment = _align("center")

        # ── Freeze panes ──────────────────────────────────────────
        ws.freeze_panes = "C3"   # freeze cột A,B và row 1,2

        # ── Print settings ────────────────────────────────────────
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1

    wb.save(output_path)
