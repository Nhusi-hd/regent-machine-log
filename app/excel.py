from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Color constants (match Regent template) ──────────────────────
YELLOW      = "FFD700"   # nhập tay
BLUE_INPUT  = "00B0F0"   # công thức (xanh nhạt như template)
NAVY        = "1A2744"   # header nền
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F2F2F2"
ORANGE      = "E8542A"   # Regent accent
GREEN_HDR   = "00B050"   # section header xanh lá (như template gốc)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _write(ws, row, col, value, fill_hex=None, bold=False,
           font_color="000000", align="left", wrap=False, formula=False):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_hex:
        cell.fill = _fill(fill_hex)
    cell.font  = _font(bold=bold, color=font_color if not formula else "0070C0")
    cell.border = _border()
    cell.alignment = _center(wrap) if align == "center" else _left(wrap)
    return cell


def build_excel(entry, computed: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Log"

    # ── Column widths ────────────────────────────────────────────
    col_widths = {1: 28, 2: 20}
    for c in col_widths:
        ws.column_dimensions[get_column_letter(c)].width = col_widths[c]
    # Date columns (col 3+) = 12 each
    ws.column_dimensions[get_column_letter(3)].width = 16

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # ── Title row ────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    title = ws["A1"]
    title.value = "DAILY MACHINE PERFORMANCE LOG — REGENT GARMENT FACTORY"
    title.fill  = _fill(NAVY)
    title.font  = _font(bold=True, color=WHITE, size=12)
    title.alignment = _center()
    title.border = _border()

    # ── Sub-header: project + date ────────────────────────────────
    ws.merge_cells("A2:C2")
    sub = ws["A2"]
    sub.value = f"Project: {entry.project}  |  Date: {entry.date}  |  MO/Color: {entry.mo_color}  |  Machine: {entry.machine_line}"
    sub.fill  = _fill(GRAY_LIGHT)
    sub.font  = _font(bold=False, size=9, color="444444")
    sub.alignment = _left()
    sub.border = _border()

    # ── Section header helper ─────────────────────────────────────
    def section_header(row, text):
        ws.merge_cells(f"A{row}:C{row}")
        ws.row_dimensions[row].height = 18
        c = ws[f"A{row}"]
        c.value = text
        c.fill  = _fill(NAVY)
        c.font  = _font(bold=True, color=WHITE, size=10)
        c.alignment = _left()
        c.border = _border()

    # ── Field row helper ──────────────────────────────────────────
    # col A = label, col B = value (Yellow=manual / Blue=formula), col C = note
    def field(row, label, value, is_formula=False, note=""):
        ws.row_dimensions[row].height = 16
        fill_hex = BLUE_INPUT if is_formula else YELLOW
        _write(ws, row, 1, label,  GRAY_LIGHT, bold=False, align="left")
        _write(ws, row, 2, value,  fill_hex,   bold=is_formula,
               font_color="0070C0" if is_formula else "000000", align="center")
        _write(ws, row, 3, note,   WHITE, bold=False, align="left",
               font_color="888888")
        # formula marker
        if is_formula:
            ws.cell(row=row, column=2).font = _font(bold=True, color="0070C0", size=10)

    # ══════════════════════════════════════════════════════════════
    # SECTION 1 – Thông tin chung
    # ══════════════════════════════════════════════════════════════
    R = 4
    section_header(R, "1.  THÔNG TIN CHUNG"); R += 1

    field(R, "SAM (testing)",              entry.sam,                    note="phút/sản phẩm"); R += 1
    field(R, "Standard Working Time",      entry.standard_working_time,  note="giờ / ngày"); R += 1
    field(R, "HC vận hành máy",            entry.hc_operators,           note="người"); R += 1

    # ══════════════════════════════════════════════════════════════
    # SECTION 2 – Output / Sản lượng
    # ══════════════════════════════════════════════════════════════
    section_header(R, "2.  OUTPUT / SẢN LƯỢNG"); R += 1

    # Formula cells (Blue)
    field(R, "Target output / day (pcs)",
          computed["target_output_day"],
          is_formula=True,
          note=f"= (SWT×60)÷SAM = ({entry.standard_working_time}×60)÷{entry.sam}"); R += 1

    field(R, "Auto target output / h / HC",
          computed["auto_target_h_hc"],
          is_formula=True,
          note=f"= 60÷SAM÷HC = 60÷{entry.sam}÷{entry.hc_operators}"); R += 1

    # Manual inputs (Yellow)
    field(R, "Actual output / day / HC (pcs)", entry.actual_output_day_hc);  R += 1
    field(R, "Actual output / day / MC (pcs)", entry.actual_output_day_mc);  R += 1
    field(R, "Auto output / h / HC (pcs)",     entry.auto_output_h_hc);      R += 1
    field(R, "Actual output / h / MC (pcs)",   entry.actual_output_h_mc);    R += 1
    field(R, "Manual output / day / HC (pcs)", entry.manual_output_day_hc);  R += 1
    field(R, "Manual output / h / HC (pcs)",   entry.manual_output_h_hc);    R += 1

    # KPI formulas
    field(R, "EFF (%)",
          f"{computed['eff_pct']}%",
          is_formula=True,
          note="= Actual HC ÷ Target/day × 100"); R += 1

    field(R, "Defect rate (%)",
          f"{computed['defect_rate_pct']}%",
          is_formula=True,
          note="= Total defect ÷ Actual MC × 100"); R += 1

    # ══════════════════════════════════════════════════════════════
    # SECTION 3 – Thời gian máy
    # ══════════════════════════════════════════════════════════════
    section_header(R, "3.  THỜI GIAN MÁY  (phút)"); R += 1

    field(R, "Machine working time (min)", entry.machine_working_time);  R += 1
    field(R, "Change-over time (min)",     entry.changeover_time);       R += 1
    field(R, "Breakdown time (min)",       entry.breakdown_time);        R += 1
    field(R, "Idle time (min)",            entry.idle_time);             R += 1

    field(R, "Total time (min)",
          computed["total_time_min"],
          is_formula=True,
          note="= Working + Change-over + Breakdown + Idle"); R += 1

    field(R, "MC Utilization (%)",
          f"{computed['mc_utilization_pct']}%",
          is_formula=True,
          note="= Machine working time ÷ Total time × 100"); R += 1

    # ══════════════════════════════════════════════════════════════
    # SECTION 4 – Defect
    # ══════════════════════════════════════════════════════════════
    section_header(R, "4.  DEFECT  (pcs)"); R += 1

    # Defect table header
    ws.row_dimensions[R].height = 16
    for col, (text, w) in enumerate(
        [("Tên lỗi", 28), ("Số lượng (pcs)", 16), ("% / MC output", 16)], start=1
    ):
        c = ws.cell(row=R, column=col, value=text)
        c.fill = _fill(GREEN_HDR)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.border = _border()
        c.alignment = _center()
    R += 1

    defect_start_row = R
    total_mc = entry.actual_output_day_mc

    for d in entry.defects:
        pct = round(d.qty / total_mc * 100, 2) if total_mc > 0 else 0
        ws.row_dimensions[R].height = 15
        _write(ws, R, 1, d.name, YELLOW,      align="left")
        _write(ws, R, 2, d.qty,  YELLOW,      align="center")
        _write(ws, R, 3, f"{pct}%", BLUE_INPUT, align="center",
               bold=True, font_color="0070C0", formula=True)
        R += 1

    # Total defect row
    ws.row_dimensions[R].height = 17
    _write(ws, R, 1, "Total defect", BLUE_INPUT, bold=True,
           font_color="0070C0", align="left")
    _write(ws, R, 2, computed["total_defect"], BLUE_INPUT, bold=True,
           font_color="0070C0", align="center")
    _write(ws, R, 3, "", BLUE_INPUT)
    R += 2

    # ══════════════════════════════════════════════════════════════
    # LEGEND
    # ══════════════════════════════════════════════════════════════
    ws.merge_cells(f"A{R}:C{R}")
    lg = ws[f"A{R}"]
    lg.value = "⬛ Màu vàng = Nhập tay     🔵 Màu xanh = Công thức tự tính"
    lg.fill  = _fill(GRAY_LIGHT)
    lg.font  = _font(size=8, color="666666")
    lg.alignment = _left()
    lg.border = _border()
    R += 1

    # Footer
    ws.merge_cells(f"A{R}:C{R}")
    ft = ws[f"A{R}"]
    ft.value = "Regent Garment Factory  |  Crystal International Group  |  Innovation Engineering Dept."
    ft.fill  = _fill(NAVY)
    ft.font  = _font(bold=False, color=WHITE, size=8)
    ft.alignment = _center()
    ft.border = _border()

    # ── Print settings ────────────────────────────────────────────
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    wb.save(output_path)
